import tempfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from DesktopAssistant.database.excel_manager import ExcelManager


class TestExcelManager:
    def setup_method(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_dir = Path(self.temp_dir.name)
        self.manager = ExcelManager(base_dir=self.base_dir)

    def teardown_method(self):
        self.temp_dir.cleanup()

    def test_creates_workbook_with_headers(self):
        assert self.manager.excel_path.exists()
        records = self.manager.list_records()
        assert records == []

    def test_add_and_get_record(self):
        record_id = self.manager.add_record(
            tags="工作 账号",
            content_type="text",
            text_content="邮箱: example@company.com",
            image_path="",
            note="测试备注",
        )

        record = self.manager.get_record(record_id)
        assert record is not None
        assert record["id"] == record_id
        assert record["tags"] == "工作 账号"
        assert record["content_type"] == "text"

    def test_search_and_filters(self):
        self.manager.add_record(
            tags="工作 邮箱",
            content_type="text",
            text_content="mail.company.com",
        )
        self.manager.add_record(
            tags="项目 截图",
            content_type="screenshot",
            text_content="UI稿",
            image_path="images/a.png",
        )

        assert len(self.manager.list_records(search_query="mail")) == 1
        assert len(self.manager.list_records(tag="工作")) == 1
        assert len(self.manager.list_records(content_type="screenshot")) == 1

    def test_delete_record(self):
        record_id = self.manager.add_record(
            tags="待删",
            content_type="text",
            text_content="to delete",
        )
        deleted = self.manager.delete_record(record_id)

        assert deleted is True
        assert self.manager.get_record(record_id) is None

    def test_export_embeds_image_instead_of_path_text(self):
        image_file = self.base_dir / "images" / "sample.png"
        image_file.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (80, 50), color="orange").save(image_file)

        self.manager.add_record(
            tags="截图 测试",
            content_type="screenshot",
            text_content="contains image",
            image_path="images/sample.png",
            note="for export",
        )

        export_path = self.base_dir / "exports" / "export.xlsx"
        self.manager.export_to_path(export_path)

        workbook = load_workbook(export_path)
        sheet = workbook[self.manager.SHEET_NAME]

        assert sheet.cell(row=1, column=6).value == "图片"
        assert sheet.cell(row=2, column=6).value in ("", None)
        assert len(sheet._images) == 1
