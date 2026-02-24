import tempfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from DesktopAssistant.app_core import AssistantCore


class TestAssistantCore:
    def setup_method(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_dir = Path(self.temp_dir.name)
        self.core = AssistantCore(base_dir=self.base_dir)

    def teardown_method(self):
        self.temp_dir.cleanup()

    def test_save_text_record(self):
        record = self.core.save_entry(text="hello", tags="工作 账号")

        assert record["id"] > 0
        assert record["content_type"] == "text"
        assert record["text_content"] == "hello"

    def test_save_mixed_record_with_image(self):
        image = Image.new("RGB", (20, 20), color="green")
        record = self.core.save_entry(text="doc", tags="截图", image=image)

        assert record["content_type"] == "mixed"
        assert record["image_path"].startswith("images/")
        assert self.core.image_manager.get_absolute_path(record["image_path"]).exists()

    def test_search_records(self):
        self.core.save_entry(text="mail.company.com", tags="工作 邮箱")
        self.core.save_entry(text="ui image", tags="项目 截图")

        records = self.core.list_entries(search_query="mail")
        assert len(records) == 1
        assert records[0]["tags"] == "工作 邮箱"

    def test_export_excel(self):
        self.core.save_entry(text="export me", tags="导出 测试")
        output_path = self.base_dir / "exports" / "records_export.xlsx"

        exported = self.core.export_excel(output_path)

        assert exported == output_path
        assert output_path.exists()

    def test_save_entry_applies_tag_alias_map(self):
        self.core.config.set("tag_alias_map", {"1": "工作", "2": "账号"})
        record = self.core.save_entry(text="mapped", tags="1 2 项目 1")

        assert record["tags"] == "工作 账号 项目"

    def test_export_excel_with_tag_filter(self):
        self.core.save_entry(text="mail", tags="工作 邮箱")
        self.core.save_entry(text="ui", tags="项目 截图")
        output_path = self.base_dir / "exports" / "work_only.xlsx"

        self.core.export_excel(output_path, tag="工作")

        workbook = load_workbook(output_path)
        sheet = workbook["records"]
        assert sheet.max_row == 2
        assert sheet.cell(row=2, column=3).value == "工作 邮箱"

    def test_update_entry_applies_alias_and_updates_content_type(self):
        self.core.config.set("tag_alias_map", {"1": "工作"})
        image = Image.new("RGB", (10, 10), color="blue")
        record = self.core.save_entry(text="", tags="项目", image=image)

        updated = self.core.update_entry(
            record["id"],
            tags="1",
            text_content="补充说明",
            note="已更新",
        )

        assert updated is True
        current = self.core.get_entry(record["id"])
        assert current is not None
        assert current["tags"] == "工作"
        assert current["content_type"] == "mixed"
        assert current["note"] == "已更新"

    def test_update_entry_replaces_image_and_removes_old_file(self):
        old_image = Image.new("RGB", (10, 10), color="blue")
        record = self.core.save_entry(text="", tags="项目", image=old_image)
        old_path = self.core.image_manager.get_absolute_path(record["image_path"])
        assert old_path.exists()

        new_image = Image.new("RGB", (12, 12), color="green")
        updated = self.core.update_entry(record["id"], image=new_image)

        assert updated is True
        current = self.core.get_entry(record["id"])
        assert current is not None
        assert current["image_path"].startswith("images/")
        assert current["image_path"] != record["image_path"]
        assert self.core.image_manager.get_absolute_path(current["image_path"]).exists()
        assert old_path.exists() is False
