"""Excel-based storage backend."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment


@dataclass(frozen=True)
class Columns:
    ID: str = "ID"
    CREATED_AT: str = "创建时间"
    TAGS: str = "标签"
    CONTENT_TYPE: str = "内容类型"
    TEXT_CONTENT: str = "文本内容"
    IMAGE_PATH: str = "图片路径"
    NOTE: str = "备注"

    @classmethod
    def ordered(cls) -> list[str]:
        return [
            cls.ID,
            cls.CREATED_AT,
            cls.TAGS,
            cls.CONTENT_TYPE,
            cls.TEXT_CONTENT,
            cls.IMAGE_PATH,
            cls.NOTE,
        ]


class ExcelManager:
    """Persist records into a local Excel workbook."""

    SHEET_NAME = "records"

    def __init__(self, base_dir: Optional[Path] = None, excel_path: Optional[Path] = None) -> None:
        if excel_path is not None:
            self.excel_path = Path(excel_path)
            self.base_dir = self.excel_path.parent
        else:
            self.base_dir = Path(base_dir) if base_dir else Path.home() / "DesktopAssistant"
            self.excel_path = self.base_dir / "data.xlsx"

        self.images_dir = self.base_dir / "images"
        self.backups_dir = self.base_dir / "backups"
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.images_dir.mkdir(parents=True, exist_ok=True)
        self.backups_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if self.excel_path.exists():
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.SHEET_NAME
        sheet.append(Columns.ordered())
        workbook.save(self.excel_path)

    def _open_sheet(self):
        workbook = load_workbook(self.excel_path)
        if self.SHEET_NAME not in workbook.sheetnames:
            sheet = workbook.active
            sheet.title = self.SHEET_NAME
            sheet.append(Columns.ordered())
        sheet = workbook[self.SHEET_NAME]
        return workbook, sheet

    @staticmethod
    def _row_to_record(row_values: list[Any]) -> dict[str, Any]:
        created_at = row_values[1]
        if isinstance(created_at, datetime):
            created_at_value = created_at.isoformat(timespec="seconds")
        else:
            created_at_value = str(created_at or "")

        return {
            "id": int(row_values[0]),
            "created_at": created_at_value,
            "tags": str(row_values[2] or ""),
            "content_type": str(row_values[3] or ""),
            "text_content": str(row_values[4] or ""),
            "image_path": str(row_values[5] or ""),
            "note": str(row_values[6] or ""),
        }

    def _next_id(self, sheet) -> int:
        max_id = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell_value = row[0]
            if isinstance(cell_value, int):
                max_id = max(max_id, cell_value)
        return max_id + 1

    def add_record(
        self,
        *,
        tags: str,
        content_type: str,
        text_content: str = "",
        image_path: str = "",
        note: str = "",
    ) -> int:
        workbook, sheet = self._open_sheet()
        record_id = self._next_id(sheet)
        sheet.append(
            [
                record_id,
                datetime.now(),
                tags,
                content_type,
                text_content,
                image_path,
                note,
            ]
        )
        workbook.save(self.excel_path)
        return record_id

    def list_records(
        self,
        *,
        search_query: str = "",
        tag: Optional[str] = None,
        content_type: Optional[str] = None,
        limit: Optional[int] = None,
        offset: int = 0,
    ) -> list[dict[str, Any]]:
        _, sheet = self._open_sheet()
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        records = [self._row_to_record(list(row)) for row in rows if row and row[0] is not None]

        if search_query:
            query = search_query.lower()
            records = [
                item
                for item in records
                if query in item["tags"].lower()
                or query in item["text_content"].lower()
                or query in item["note"].lower()
            ]

        if tag:
            tag_lower = tag.lower()
            records = [item for item in records if tag_lower in item["tags"].lower().split()]

        if content_type:
            records = [item for item in records if item["content_type"] == content_type]

        records.sort(key=lambda item: item["id"], reverse=True)
        if offset:
            records = records[offset:]
        if limit is not None:
            records = records[:limit]
        return records

    def get_record(self, record_id: int) -> Optional[dict[str, Any]]:
        _, sheet = self._open_sheet()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == record_id:
                return self._row_to_record(list(row))
        return None

    def delete_record(self, record_id: int) -> bool:
        workbook, sheet = self._open_sheet()
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] == record_id:
                sheet.delete_rows(index, 1)
                workbook.save(self.excel_path)
                return True
        return False

    def update_record(
        self,
        record_id: int,
        *,
        tags: Optional[str] = None,
        content_type: Optional[str] = None,
        text_content: Optional[str] = None,
        image_path: Optional[str] = None,
        note: Optional[str] = None,
    ) -> bool:
        workbook, sheet = self._open_sheet()
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value == record_id:
                if tags is not None:
                    sheet.cell(row=row_index, column=3, value=tags)
                if content_type is not None:
                    sheet.cell(row=row_index, column=4, value=content_type)
                if text_content is not None:
                    sheet.cell(row=row_index, column=5, value=text_content)
                if image_path is not None:
                    sheet.cell(row=row_index, column=6, value=image_path)
                if note is not None:
                    sheet.cell(row=row_index, column=7, value=note)
                workbook.save(self.excel_path)
                return True
        return False

    def available_tags(self) -> list[str]:
        tags: set[str] = set()
        for record in self.list_records():
            tags.update(tag for tag in record["tags"].split() if tag)
        return sorted(tags)

    def export_to_path(
        self,
        output_path: Path,
        records: Optional[list[dict[str, Any]]] = None,
    ) -> Path:
        """Export records to a target workbook with embedded images."""
        target = Path(output_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        export_records = records if records is not None else self.list_records()

        export_book = Workbook()
        export_sheet = export_book.active
        export_sheet.title = self.SHEET_NAME
        export_sheet.append(
            [
                Columns.ID,
                Columns.CREATED_AT,
                Columns.TAGS,
                Columns.CONTENT_TYPE,
                Columns.TEXT_CONTENT,
                "图片",
                Columns.NOTE,
            ]
        )
        export_sheet.freeze_panes = "A2"
        export_sheet.column_dimensions["E"].width = 48
        export_sheet.column_dimensions["F"].width = 36

        for record in export_records:
            record_id = record.get("id")
            created_at = record.get("created_at", "")
            tags = record.get("tags", "")
            content_type = record.get("content_type", "")
            text_content = record.get("text_content", "")
            image_path = record.get("image_path", "")
            note = record.get("note", "")
            export_row = export_sheet.max_row + 1
            export_sheet.append([record_id, created_at, tags, content_type, text_content, None, note])
            export_sheet.cell(row=export_row, column=5).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
            export_sheet.cell(row=export_row, column=7).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

            if not image_path:
                continue

            image_file = self.base_dir / str(image_path)
            if not image_file.exists():
                continue

            try:
                excel_image = XLImage(str(image_file))
            except OSError:
                continue

            max_width, max_height = 240, 140
            if excel_image.width and excel_image.height:
                scale = min(max_width / excel_image.width, max_height / excel_image.height, 1.0)
                excel_image.width = int(excel_image.width * scale)
                excel_image.height = int(excel_image.height * scale)

            export_sheet.add_image(excel_image, f"F{export_row}")
            current_height = export_sheet.row_dimensions[export_row].height or 15
            expected_height = excel_image.height * 0.75 + 6
            export_sheet.row_dimensions[export_row].height = max(current_height, expected_height)

        export_book.save(target)
        export_book.close()
        return target
